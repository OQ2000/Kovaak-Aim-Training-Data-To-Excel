import pyautogui, sys
from pyautogui import hotkey, press, moveTo, move, click
from time import sleep
import clipboard
from pynput import keyboard
import os, os.path ,shutil, glob
import math as math
import getpass
import openpyxl 
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from datetime import date
import msvcrt
import csv
from xlsxwriter.workbook import Workbook
clear = lambda: os.system('cls')
clear()

def MoveAllFiles(srcDir, dstDir):
    def moveAllFilesinDir(srcDir, dstDir):
    # Check if both the are directories
        if os.path.isdir(srcDir) and os.path.isdir(dstDir) :
        # Iterate over all the files in source directory
            for filePath in glob.glob(srcDir + r'\*'):
            # Move each file to destination Directory
                shutil.move(filePath, dstDir);
            else:
                print("Files Have Been Moved To The Directory")    
    moveAllFilesinDir(srcDir, dstDir)
    sleep(4)

def GetDataFromCSVName(Name):
    Name = Name.replace("- Challenge -","")
    Name = Name.replace("Stats","")
    Name = Name.replace(" ", "")
    Name = Name[0:-5]
    #CloseLongStrafesInvincible2020.05.13-03.04.31
    tmplen = len(Name)
    Time = Name[tmplen-8::]
    Time = Time.replace(".",":")
    Name = Name[0:tmplen-9]
    tmplen = len(Name)
    Date = Name[tmplen-10::]
    Date = Date.replace(".","/")
    Name = Name[0:tmplen-10]
    return(Name,Date,Time)

def FormatAllExcels(wsActive):
    ws = wsActive.worksheets[0]
    NextEmpty = get_next_empty_cell(ws)
    ws.delete_rows(0,int(NextEmpty))

def GetNoOfCSVs(dir):
    listOfFiles = next(os.walk(dir))[2] #dir is your directory path as string
    NoOfFIles = len(listOfFiles)
    return(NoOfFIles, listOfFiles)

def ConvertAllCSVToExcel():
    for csvfile in glob.glob(os.path.join('.', '*.csv')):
        workbook = Workbook(csvfile[:-4] + '.xlsx')
        worksheet = workbook.add_worksheet()
        with open(csvfile, 'rt', encoding='utf8') as f:
            reader = csv.reader(f)
            for r, row in enumerate(reader):
                for c, col in enumerate(row):
                    worksheet.write(r, c, col)
        workbook.close()
    directory = os.getcwd()
    files_in_directory = os.listdir(directory)
    filtered_files = [file for file in files_in_directory if file.endswith(".csv")]
    for file in filtered_files:
        path_to_file = os.path.join(directory, file)
        os.remove(path_to_file)

def CheckIfTestExistsInMaster(wbMaster, FindSheetName):
    for i in range(len(wbMaster.sheetnames)):
        if(wbMaster.sheetnames[i] == FindSheetName):
            print("FoundSheet")
            return(True)
        print(wbMaster.sheetnames[i])
    return(False)

def GetDataFromExcelIntake(wb):
    ws = wb.worksheets[0]
    Score = float(ws['B15'].value)
    Shots = float(ws['B2'].value)
    Hits = int(ws['C2'].value)
    DamageDone = float(ws['D2'].value)
    DamagePossible = float(ws['E2'].value)
    return(Score,Shots,Hits,DamageDone,DamagePossible)

def get_next_empty_cell(ws):
    for cell in ws["A"]:
        if cell.value is None:
            print(cell.row)
            break
        else:
            Empty_Cell_Row = cell.row + 1
            # print("Empty cell At: A" + str(Empty_Cell_Row))
    return(Empty_Cell_Row)

def ConvertNumberToLetter(num):
    alphabetList = {
        0 : "A",
        1 : "B",
        2 : "C",
        3 : "D",
        4 : "E",
        5 : "F",
        6 : "G",
        7 : "H",
        8 : "I",
        9 : "J",
        10 : "K",
        11 : "L",
        12 : "M",
        13 : "N",
        14 : "O",
        15 : "P",
        16 : "Q",
        17 : "R",
        18 : "S",
        19 : "T",
        20 : "U",
        21 : "V",
        22 : "W",
        23 : "X",
        24 : "Y",
        25 : "Z",
    }
    print(num, "Converts To:", alphabetList[num])
    return(alphabetList[num])

def SizeColoumns(ws):
    for x in range(12):
        tmpLetter = ConvertNumberToLetter(x)
        ws.column_dimensions[tmpLetter].width = 20

def Main():
    path = os.getcwd()
    os.chdir(path)
    wbMaster = load_workbook('InputForAimTraining\MasterAimTrainingDataSet.xlsx')

    inputpath = path + r"\InputForAimTraining\csv's to be input"
    os.chdir(inputpath)
    ConvertAllCSVToExcel()
    os.chdir(path)
    NoOfCSVs, listOfFiles = GetNoOfCSVs(inputpath)
    # FindSheetName = "Sheet"
    # tmpResult = CheckIfTestExistsInMaster(wbMaster, FindSheetName)
    # print(tmpResult)
    for i in range(NoOfCSVs):
        NameOfFile = listOfFiles[i]
        Name, Date, Time = GetDataFromCSVName(NameOfFile)
        print("Name:",Name)
        print("Date:",Date)
        print("Time:",Time)
        #Get Data From And Place In wb
        wbIntake = load_workbook(inputpath+"\\"+NameOfFile)
        FormatAllExcels(wbIntake)
        Score,Shots,Hits,DamageDone,DamagePossible = GetDataFromExcelIntake(wbIntake)
        print("Score:",Score,"Shots:",Shots,"Hits:",Hits,"DamageDone:",DamageDone,"DamagePossible:",DamagePossible)
        if CheckIfTestExistsInMaster(wbMaster, Name) == False:
            wbMaster.create_sheet(Name)
            NameIndex = wbMaster.sheetnames.index(Name)
            wsActive = wbMaster.worksheets[NameIndex]
            wsActive['A1'].value = "Date"
            wsActive['B1'].value = "Time"
            wsActive['C1'].value = "Score"
            wsActive['D1'].value = "Shots"
            wsActive['E1'].value = "Hits"
            wsActive['F1'].value = "Accuracy"
            wsActive['G1'].value = "DamageDone"
            wsActive['H1'].value = "DamagePossible"
            alignment = Alignment(horizontal='left', vertical='center')
            for g in range(12):
                tmpLetter = ConvertNumberToLetter(g)
                col = wsActive.column_dimensions[tmpLetter]
                col.Alignment = alignment
            SizeColoumns(wsActive)
        NameIndex = wbMaster.sheetnames.index(Name)
        wsActive = wbMaster.worksheets[NameIndex]
        Next_Empty_Row = get_next_empty_cell(wsActive)
        Next_Empty_Row = str(Next_Empty_Row)
        wsActive["A"+Next_Empty_Row].value = Date
        wsActive["A"+Next_Empty_Row].number_format = "dd/mm/yyyy;"

        wsActive["B"+Next_Empty_Row].value = Time
        wsActive["B"+Next_Empty_Row].number_format = "h:mm:ss;"

        wsActive["C"+Next_Empty_Row].value = Score
        wsActive["C"+Next_Empty_Row].number_format = "0.0"

        wsActive["D"+Next_Empty_Row].value = Shots
        wsActive["D"+Next_Empty_Row].number_format = "0.0"

        wsActive["E"+Next_Empty_Row].value = Hits
        wsActive["E"+Next_Empty_Row].number_format = "0.0"

        wsActive["F"+Next_Empty_Row].value = '=E'+Next_Empty_Row+'/D'+Next_Empty_Row
        wsActive["F"+Next_Empty_Row].number_format = "0.00%"

        wsActive["G"+Next_Empty_Row].value = DamageDone
        wsActive["G"+Next_Empty_Row].number_format = "0.0"

        wsActive["H"+Next_Empty_Row].value = DamagePossible
        wsActive["H"+Next_Empty_Row].number_format = "0.0"
        
    return(wbMaster, path)

srcDir = r"C:\Program Files (x86)\Steam\steamapps\common\FPSAimTrainer\FPSAimTrainer\stats"
dstDir = r"C:\Users\Owen\Desktop\Projects\InputForAimTraining\csv's to be input"

MoveAllFiles(srcDir, dstDir)
wbMaster, path= Main()
print("Saving To:",path + r"\InputForAimTraining\MasterAimTrainingDataSet.xlsx")
wbMaster.save(filename = path + r"\InputForAimTraining\MasterAimTrainingDataSet.xlsx")

archive = r"C:\Users\Owen\Desktop\Projects\InputForAimTraining\Archived"
MoveAllFiles(dstDir,archive)